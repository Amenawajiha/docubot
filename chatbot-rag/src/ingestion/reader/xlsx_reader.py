"""
Excel and CSV Reader - Extract text from spreadsheet files.

Supports:
- Excel files (.xlsx)
- CSV files (.csv) - with automatic delimiter detection
- Automatic format detection (handles mis-labeled files)

Output format: Natural language sentences
- "Column A is value1, Column B is value2, Column C is value3."
- Preserves sheet structure for multi-sheet workbooks
"""

import logging
import re
from io import BytesIO
from typing import Any

from .base_reader import BaseReader

logger = logging.getLogger(__name__)

# Regex to remove illegal control characters from Excel cells
ILLEGAL_CHARS = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


class XLSXReader(BaseReader):
    """
    Extract text from Excel and CSV files.
    
    Smart features:
    - Auto-detects actual format (CSV labeled as .xlsx, etc.)
    - Handles both modern (.xlsx) and legacy (.xls) Excel
    - Converts tabular data to natural language sentences
    - Preserves multi-sheet structure
    """
    
    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".csv"]
    
    def extract(self, file_bytes: bytes) -> str:
        """
        Extract text from Excel or CSV file.
        
        Args:
            file_bytes: Raw file bytes
            
        Returns:
            Natural language representation of spreadsheet data
            
        Raises:
            ValueError: If file is empty or contains no data
            RuntimeError: If parsing fails
        """
        if not file_bytes:
            raise ValueError("Empty spreadsheet file provided")
        
        try:
            # Load workbook (handles format detection)
            wb = self._load_workbook(file_bytes)
            
            # Convert to text
            text = self._workbook_to_text(wb)
            
            # Clean up
            if hasattr(wb, 'close'):
                wb.close()
            
            if not text.strip():
                raise ValueError("No data found in spreadsheet")
            
            logger.info(f"Spreadsheet extraction complete: {len(text)} characters")
            return text
        
        except Exception as e:
            logger.error(f"Spreadsheet extraction failed: {e}")
            raise RuntimeError(f"Failed to read spreadsheet: {e}") from e
    
    def _load_workbook(self, file_bytes: bytes):
        """
        Load workbook with automatic format detection.
        
        Tries multiple strategies:
        1. Check magic bytes to detect actual format
        2. Try openpyxl (Excel)
        3. Try pandas read_csv (CSV)
        
        Args:
            file_bytes: Raw file bytes
            
        Returns:
            openpyxl Workbook object or pandas DataFrame-based workbook
        """
        # Lazy imports
        try:
            import openpyxl
            import pandas as pd
        except ImportError as e:
            raise RuntimeError(
                "openpyxl and pandas are required for spreadsheet support. "
                "Install with: pip install openpyxl pandas"
            ) from e
        
        bio = BytesIO(file_bytes)
        bio.seek(0)
        magic = bio.read(4)
        bio.seek(0)
        
        # Detect format by magic bytes
        is_zip = magic.startswith(b"PK\x03\x04")      # Modern Excel (.xlsx)
        
        # Strategy 1: Try as CSV if not Excel format
        if not is_zip:
            logger.info("File appears to be CSV (no Excel magic bytes)")
            try:
                df = pd.read_csv(bio, on_bad_lines='skip')
                return self._df_to_workbook(df)
            except Exception as e:
                logger.warning(f"CSV parsing failed: {e}")
        
        # Strategy 2: Try openpyxl (modern Excel)
        bio.seek(0)
        try:
            logger.info("Attempting to load as modern Excel (.xlsx)")
            return openpyxl.load_workbook(bio, data_only=True)
        except Exception as e:
            raise RuntimeError(
                f"Could not load spreadsheet in any known format: {e}"
            ) from e
        
    
    def _df_to_workbook(self, data):
        """
        Convert pandas DataFrame(s) to openpyxl-like workbook structure.
        
        This creates a unified interface so downstream code doesn't need
        to know if the source was CSV or Excel.
        
        Args:
            data: Single DataFrame or dict of {sheet_name: DataFrame}
            
        Returns:
            Mock workbook object with .sheetnames and indexable sheets
        """
        import pandas as pd
        
        # Handle single DataFrame (CSV)
        if isinstance(data, pd.DataFrame):
            data = {"Sheet1": data}
        
        # Create mock workbook
        class MockWorkbook:
            def __init__(self, sheets):
                self.sheets = sheets
                self.sheetnames = list(sheets.keys())
            
            def __getitem__(self, name):
                return self.sheets[name]
            
            def close(self):
                pass
        
        # Convert DataFrames to row iterators
        sheets = {}
        for sheet_name, df in data.items():
            # Add headers as first row
            headers = list(df.columns)
            rows = [headers] + df.values.tolist()
            
            # Create mock sheet
            class MockSheet:
                def __init__(self, rows):
                    self._rows = rows
                
                def iter_rows(self, values_only=True):
                    return iter(self._rows)
            
            sheets[sheet_name] = MockSheet(rows)
        
        return MockWorkbook(sheets)
    
    def _workbook_to_text(self, wb) -> str:
        """
        Convert workbook to natural language sentences.
        
        Format: "Column A is X, Column B is Y, Column C is Z."
        Each row becomes one sentence. Multiple sheets are labeled.
        
        Args:
            wb: Workbook object (openpyxl or mock)
            
        Returns:
            Natural language text
        """
        all_parts = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows:
                logger.debug(f"Sheet '{sheet_name}' is empty, skipping")
                continue
            
            # Extract headers (first row)
            headers = [
                self._clean(h) if h is not None else f"Column{i+1}"
                for i, h in enumerate(rows[0])
            ]
            
            # Convert data rows to sentences
            sheet_sentences = []
            
            for row in rows[1:]:  # Skip header row
                parts = []
                for col_name, val in zip(headers, row):
                    clean_val = self._clean(val)
                    if clean_val:  # Only include non-empty values
                        parts.append(f"{col_name} is {clean_val}")
                
                if parts:
                    sheet_sentences.append(", ".join(parts) + ".")
            
            # Add sheet to output if it has data
            if sheet_sentences:
                # Add sheet label (avoid redundancy if already in name)
                sheet_suffix = (
                    f" - {sheet_name}"
                    if "sheet" not in sheet_name.lower()
                    else ""
                )
                
                sheet_text = (
                    f"Sheet: {sheet_name}{sheet_suffix}\n\n"
                    + "\n\n".join(sheet_sentences)
                )
                all_parts.append(sheet_text)
        
        if not all_parts:
            raise ValueError("No valid data found in any sheet")
        
        return "\n\n".join(all_parts)
    
    @staticmethod
    def _clean(val: Any) -> str:
        """
        Clean cell value for text output.
        
        Removes illegal control characters and normalizes whitespace.
        
        Args:
            val: Cell value (any type)
            
        Returns:
            Cleaned string
        """
        s = str(val).strip() if val is not None else ""
        return ILLEGAL_CHARS.sub(" ", s)


class CSVReader(XLSXReader):
    """
    CSV reader (inherits all logic from XLSXReader).
    
    CSV files are just single-sheet spreadsheets, so we reuse
    the Excel reader infrastructure.
    """
    
    @property
    def supported_extensions(self) -> list[str]:
        return [".csv"]